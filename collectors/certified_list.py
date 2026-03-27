"""健康経営優良法人 認定企業リスト収集

経産省「ACTION!健康経営」ポータルで公開されているExcelファイルから
認定企業一覧を取得し、認定企業自体と「まだ認定されていない同業種企業」を
ターゲットとして活用する。
"""

import io
import re
import tempfile
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

from collectors.base_collector import BaseCollector
from config.settings import PROJECT_ROOT, REQUEST_HEADERS, REQUEST_TIMEOUT
from sheets.models import Company


# ダウンロード済みファイルのキャッシュディレクトリ
CACHE_DIR = PROJECT_ROOT / "data" / "certified_list"


class CertifiedListCollector(BaseCollector):
    """健康経営優良法人の認定企業リストを収集"""

    LIST_PAGE_URL = "https://kenko-keiei.jp/houjin_list/"

    @property
    def source_name(self) -> str:
        return "健康経営優良法人"

    def __init__(self, category: str = "large"):
        """
        Args:
            category: "large"=大規模法人部門, "small"=中小規模法人部門, "brand"=健康経営銘柄
        """
        self.category = category
        self.session = requests.Session()
        self.session.headers.update(REQUEST_HEADERS)
        CACHE_DIR.mkdir(parents=True, exist_ok=True)

    def collect(self) -> list[Company]:
        """認定企業リストを収集"""
        # 1. Excelファイルをダウンロード or キャッシュから読み込み
        excel_data = self._download_excel()
        if excel_data is None:
            # キャッシュから読み込み
            excel_data = self._load_cache()

        if excel_data is None:
            print("[CertifiedList] Excelファイルの取得に失敗")
            return []

        # 2. Excelをパース
        return self._parse_excel(excel_data)

    def _download_excel(self) -> bytes | None:
        """ポータルサイトからExcelファイルのURLを取得してダウンロード"""
        try:
            response = self.session.get(self.LIST_PAGE_URL, timeout=REQUEST_TIMEOUT)
            if response.status_code != 200:
                return None
            response.encoding = response.apparent_encoding or "utf-8"
        except requests.RequestException as e:
            print(f"[CertifiedList] ページ取得エラー: {e}")
            return None

        soup = BeautifulSoup(response.text, "lxml")

        # Excelファイルへのリンクを検索
        keyword_map = {
            "large": ["大規模法人部門", "大規模"],
            "small": ["中小規模法人部門", "中小規模"],
            "brand": ["健康経営銘柄"],
        }
        keywords = keyword_map.get(self.category, ["大規模法人部門"])

        excel_url = None
        for link in soup.find_all("a", href=True):
            href = link.get("href", "")
            text = link.get_text(strip=True)
            # Excel系の拡張子を持つリンク
            if any(ext in href.lower() for ext in [".xlsx", ".xls", ".csv"]):
                if any(kw in text for kw in keywords):
                    excel_url = href
                    break

        if not excel_url:
            # hrefにキーワードが含まれるパターンも試す
            for link in soup.find_all("a", href=True):
                href = link.get("href", "")
                if any(ext in href.lower() for ext in [".xlsx", ".xls"]):
                    if any(kw in href for kw in keywords):
                        excel_url = href
                        break

        if not excel_url:
            print(f"[CertifiedList] {self.category}のExcelリンクが見つかりません")
            return None

        if not excel_url.startswith("http"):
            excel_url = urljoin(self.LIST_PAGE_URL, excel_url)

        # ダウンロード
        try:
            resp = self.session.get(excel_url, timeout=30)
            if resp.status_code == 200:
                # キャッシュに保存
                cache_file = CACHE_DIR / f"{self.category}_latest.xlsx"
                cache_file.write_bytes(resp.content)
                print(f"[CertifiedList] ダウンロード完了: {excel_url}")
                return resp.content
        except requests.RequestException as e:
            print(f"[CertifiedList] ダウンロードエラー: {e}")

        return None

    def _load_cache(self) -> bytes | None:
        """キャッシュからExcelファイルを読み込み"""
        cache_file = CACHE_DIR / f"{self.category}_latest.xlsx"
        if cache_file.exists():
            print(f"[CertifiedList] キャッシュから読み込み: {cache_file}")
            return cache_file.read_bytes()
        return None

    def _parse_excel(self, data: bytes) -> list[Company]:
        """ExcelファイルをパースしてCompanyリストに変換"""
        try:
            import openpyxl
        except ImportError:
            print("[CertifiedList] openpyxlが未インストールです")
            return self._parse_excel_fallback(data)

        companies = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active

            # ヘッダー行を探す
            header_row = None
            name_col = None
            industry_col = None

            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=False), 1):
                for cell in row:
                    val = str(cell.value or "").strip()
                    if val in ("法人名", "企業名", "法人名称", "名称"):
                        header_row = row_idx
                        name_col = cell.column
                    elif val in ("業種", "業種名"):
                        industry_col = cell.column

                if header_row:
                    break

            if not header_row or not name_col:
                # ヘッダーが見つからない場合、A列=No, B列=法人名と仮定
                header_row = 1
                name_col = 2

            # データ行を読み込み
            for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
                name_cell = row[name_col - 1] if name_col - 1 < len(row) else None
                company_name = str(name_cell.value or "").strip() if name_cell else ""

                if not company_name or company_name in ("", "None"):
                    continue

                industry = ""
                if industry_col and industry_col - 1 < len(row):
                    industry = str(row[industry_col - 1].value or "").strip()

                category_label = {
                    "large": "大規模法人部門",
                    "small": "中小規模法人部門",
                    "brand": "健康経営銘柄",
                }.get(self.category, self.category)

                companies.append(
                    Company(
                        name=company_name,
                        url="",
                        source=f"{self.source_name}（{category_label}）",
                        article_title=f"健康経営優良法人認定 - {category_label}",
                        article_url=self.LIST_PAGE_URL,
                        status="新規",
                        memo=f"業種: {industry}" if industry else "",
                    )
                )

            wb.close()
        except Exception as e:
            print(f"[CertifiedList] Excelパースエラー: {e}")
            return []

        return companies

    def _parse_excel_fallback(self, data: bytes) -> list[Company]:
        """openpyxlが無い場合のフォールバック（pandas使用）"""
        try:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(data))

            # 法人名カラムを探す
            name_col = None
            for col in df.columns:
                if any(kw in str(col) for kw in ["法人名", "企業名", "名称"]):
                    name_col = col
                    break

            if not name_col:
                # 2列目を法人名と仮定
                name_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]

            industry_col = None
            for col in df.columns:
                if "業種" in str(col):
                    industry_col = col
                    break

            companies = []
            for _, row in df.iterrows():
                name = str(row[name_col]).strip()
                if not name or name == "nan":
                    continue

                industry = str(row[industry_col]).strip() if industry_col else ""
                if industry == "nan":
                    industry = ""

                category_label = {
                    "large": "大規模法人部門",
                    "small": "中小規模法人部門",
                    "brand": "健康経営銘柄",
                }.get(self.category, self.category)

                companies.append(
                    Company(
                        name=name,
                        url="",
                        source=f"{self.source_name}（{category_label}）",
                        article_title=f"健康経営優良法人認定 - {category_label}",
                        article_url=self.LIST_PAGE_URL,
                        status="新規",
                        memo=f"業種: {industry}" if industry else "",
                    )
                )

            return companies
        except Exception as e:
            print(f"[CertifiedList] pandasフォールバックエラー: {e}")
            return []

    def load_from_file(self, file_path: str | Path) -> list[Company]:
        """ローカルのExcelファイルから直接読み込み"""
        path = Path(file_path)
        if not path.exists():
            print(f"[CertifiedList] ファイルが見つかりません: {path}")
            return []
        return self._parse_excel(path.read_bytes())
